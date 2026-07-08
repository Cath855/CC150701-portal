import streamlit as st
import json, base64
import html as _html
import urllib.parse, urllib.request
import requests
import streamlit.components.v1 as components
from datetime import datetime, timezone, timedelta
from pathlib import Path
import openpyxl
import pandas as pd

st.set_page_config(page_title="CC150701 · Portal de Documentos", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #F2F3F7; }
.header {
    background: #003087;
    padding: 1.3rem 2rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    color: white;
}
.header h1 { margin: 0; font-size: 1.4rem; font-weight: 700; }
.header p  { margin: 0.2rem 0 0; opacity: 0.65; font-size: 0.82rem; }
.doc-row {
    background: white;
    border-radius: 10px;
    border: 1px solid #E2E2E8;
    padding: 0.85rem 1.1rem;
    margin-bottom: 0.5rem;
    display: flex; align-items: center; gap: 12px;
}
.doc-icon {
    width: 38px; height: 38px; border-radius: 7px;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.62rem; font-weight: 700; flex-shrink: 0;
}
.icon-xlsx { background: #E8F6EE; color: #1A7A45; }
.icon-pdf  { background: #FEE9E9; color: #C0392B; }
.icon-docx { background: #E8F0FE; color: #1A56DB; }
.icon-pptx { background: #FEF0E8; color: #C0550F; }
.icon-file { background: #EDEAE4; color: #6B6760; }
.doc-name  { font-weight: 600; font-size: 0.9rem; color: #1A1814; }
.doc-meta  { font-size: 0.73rem; color: #888; margin-top: 2px; }
.stDownloadButton > button {
    background: #003087 !important; color: white !important;
    border: none !important; border-radius: 7px !important;
    font-size: 0.78rem !important; padding: 0.3rem 0.85rem !important;
    width: 100% !important;
}
.stDownloadButton > button:hover { background: #00206b !important; }
div[data-testid="stExpander"] {
    background: white !important;
    border-radius: 10px !important;
    border: 1px solid #E2E2E8 !important;
    margin-bottom: 0.5rem !important;
}
section[data-testid="stExpander"] > div { padding: 0.5rem 0 !important; }

/* ── Visor tipo hoja de Excel ─────────────────────────────── */
.xl-wrap {
    overflow: auto; max-height: 540px;
    border: 1px solid #b7b7b7; border-radius: 6px; background: white;
}
.xl-table {
    border-collapse: collapse; white-space: nowrap;
    font-family: 'Segoe UI', Calibri, Arial, sans-serif; font-size: 0.8rem;
}
.xl-table th, .xl-table td {
    border: 1px solid #d9d9d9; padding: 4px 9px; text-align: left;
}
.xl-table tbody td { background: white; color: #222; }
.xl-corner, .xl-col, .xl-rownum {
    background: #f3f3f3; color: #666; font-weight: 600; text-align: center;
}
.xl-col    { position: sticky; top: 0; z-index: 2; }
.xl-corner { position: sticky; top: 0; left: 0; z-index: 3; }
.xl-rownum { position: sticky; left: 0; z-index: 1; min-width: 38px; }
tr.xl-head td { background: #eaf1fb !important; font-weight: 600; color: #1a1a1a; }
tr.xl-head td.xl-rownum { background: #f3f3f3 !important; color: #666; }
</style>
""", unsafe_allow_html=True)

DOCS_DIR = Path("docs")
META_FILE = DOCS_DIR / "meta.json"
DOCS_DIR.mkdir(exist_ok=True)

ADMIN_PASS = "CNC2025"   # clave para el panel de administradora (subir/eliminar)

ICONS = {
    "xlsx": ("XLSX","icon-xlsx"), "xls": ("XLS","icon-xlsx"),
    "pdf":  ("PDF", "icon-pdf"),
    "docx": ("DOC", "icon-docx"), "doc": ("DOC","icon-docx"),
    "pptx": ("PPT", "icon-pptx"), "ppt": ("PPT","icon-pptx"),
    "csv":  ("CSV", "icon-file"), "txt": ("TXT","icon-file"),
}

def cargar_meta():
    if META_FILE.exists():
        return json.loads(META_FILE.read_text(encoding="utf-8"))
    return {}

def fmt_size(b):
    if b < 1024: return f"{b} B"
    if b < 1048576: return f"{b/1024:.0f} KB"
    return f"{b/1048576:.1f} MB"

ZONA_CO = timezone(timedelta(hours=-5))          # hora de Colombia
MESES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

def fmt_fecha(s):
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo:                             # GitHub responde en UTC
            dt = dt.astimezone(ZONA_CO)
        return f"{dt.day:02d} {MESES[dt.month - 1]} {dt.year} · {dt:%H:%M}"
    except Exception:
        return s

def get_icon(nombre):
    ext = nombre.rsplit(".",1)[-1].lower() if "." in nombre else ""
    return ICONS.get(ext, ("FILE","icon-file"))

def leer_excel(ruta):
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            data = list(ws.iter_rows(values_only=True))
            if not data: continue
            headers = [str(c) if c is not None else f"Col{i}" for i,c in enumerate(data[0])]
            rows = [list(r) for r in data[1:] if any(c is not None for c in r)]
            sheets[name] = pd.DataFrame(rows, columns=headers)
        wb.close()
        return sheets
    except Exception as e:
        return {"Error": pd.DataFrame([{"Mensaje": str(e)}])}

# ── Visor oficial de Microsoft (igual al del correo) ────────────────
# Requiere que el archivo tenga una URL publica. Como el repositorio es
# publico, cada archivo de docs/ es accesible en raw.githubusercontent.
REPO_RAW = "https://raw.githubusercontent.com/Cath855/CC150701-portal/main/docs/"

API_COMMITS = "https://api.github.com/repos/Cath855/CC150701-portal/commits"

@st.cache_data(ttl=600, show_spinner=False)
def fecha_en_github(nombre: str) -> str:
    """Fecha (ISO, UTC) del ultimo commit que toco docs/<nombre>.
    Sirve para fechar los archivos subidos directamente en GitHub."""
    params = urllib.parse.urlencode({"path": f"docs/{nombre}", "per_page": 1})
    req = urllib.request.Request(
        f"{API_COMMITS}?{params}",
        headers={"Accept": "application/vnd.github+json",
                 "User-Agent": "portal-cc150701"},
    )
    try:
        with urllib.request.urlopen(req, timeout=6) as r:
            datos = json.loads(r.read().decode("utf-8"))
        if datos:
            return datos[0]["commit"]["committer"]["date"]
    except Exception:
        pass
    return ""

@st.cache_data(ttl=600, show_spinner=False)
def disponible_en_repo(nombre: str) -> bool:
    """True si el archivo ya esta publicado en GitHub (tiene URL publica)."""
    try:
        req = urllib.request.Request(REPO_RAW + urllib.parse.quote(nombre), method="HEAD")
        with urllib.request.urlopen(req, timeout=6) as r:
            return r.status == 200
    except Exception:
        return False

# ── Escritura en GitHub (subir / eliminar desde la app) ─────────────
# Necesita un token guardado en los Secrets de Streamlit como:
#   github_token = "..."
GITHUB_CONTENTS = "https://api.github.com/repos/Cath855/CC150701-portal/contents/docs/"

def github_token() -> str:
    try:
        return st.secrets["github_token"]
    except Exception:
        return ""

def _gh_headers():
    return {
        "Authorization": f"Bearer {github_token()}",
        "Accept": "application/vnd.github+json",
        "User-Agent": "portal-cc150701",
    }

def _gh_sha(nombre: str):
    """sha actual del archivo en GitHub (hace falta para reemplazar o borrar)."""
    r = requests.get(GITHUB_CONTENTS + urllib.parse.quote(nombre),
                     headers=_gh_headers(), timeout=15)
    return r.json().get("sha") if r.status_code == 200 else None

def gh_subir(nombre: str, contenido: bytes):
    payload = {
        "message": f"Portal: subir {nombre}",
        "content": base64.b64encode(contenido).decode(),
        "branch": "main",
    }
    sha = _gh_sha(nombre)
    if sha:                       # el archivo ya existia -> se reemplaza
        payload["sha"] = sha
    r = requests.put(GITHUB_CONTENTS + urllib.parse.quote(nombre),
                     headers=_gh_headers(), json=payload, timeout=60)
    return r.status_code in (200, 201), r

def gh_eliminar(nombre: str):
    sha = _gh_sha(nombre)
    if not sha:
        return False, None
    payload = {"message": f"Portal: eliminar {nombre}", "sha": sha, "branch": "main"}
    r = requests.delete(GITHUB_CONTENTS + urllib.parse.quote(nombre),
                        headers=_gh_headers(), json=payload, timeout=60)
    return r.status_code == 200, r

def visor_office(nombre: str, alto: int = 660):
    url_archivo = REPO_RAW + urllib.parse.quote(nombre)
    src = urllib.parse.quote(url_archivo, safe="")
    components.iframe(
        f"https://view.officeapps.live.com/op/embed.aspx?src={src}",
        height=alto, scrolling=True,
    )

def _letra_columna(i):
    # 0 -> A, 1 -> B, ... 26 -> AA
    s, n = "", i + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def _celda(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        if v.is_integer():
            return str(int(v))
    return str(v)

def excel_a_html(df):
    """Dibuja el DataFrame como una hoja de Excel: cuadricula,
    letras de columna (A, B, C...), numeros de fila y encabezado."""
    cols = list(df.columns)

    letras = "<th class='xl-corner'></th>" + "".join(
        f"<th class='xl-col'>{_letra_columna(i)}</th>" for i in range(len(cols))
    )

    # La primera fila de la hoja son los nombres de columna leidos del Excel
    encabezado = "<td class='xl-rownum'>1</td>" + "".join(
        f"<td>{_html.escape('' if str(c).startswith('Col') else str(c))}</td>"
        for c in cols
    )

    filas = [f"<tr class='xl-head'>{encabezado}</tr>"]
    for n, (_, fila) in enumerate(df.iterrows(), start=2):
        celdas = f"<td class='xl-rownum'>{n}</td>" + "".join(
            f"<td>{_html.escape(_celda(v))}</td>" for v in fila
        )
        filas.append(f"<tr>{celdas}</tr>")

    return (
        "<div class='xl-wrap'><table class='xl-table'>"
        f"<thead><tr>{letras}</tr></thead>"
        f"<tbody>{''.join(filas)}</tbody>"
        "</table></div>"
    )

def mostrar_archivo(ruta: Path):
    ext = ruta.suffix.lower()
    if ext in (".xlsx", ".xls"):
        # Vista fiel (colores, celdas combinadas, pestanas de hojas) usando
        # el visor de Microsoft. Solo funciona con archivos ya publicados.
        if disponible_en_repo(ruta.name):
            visor_office(ruta.name)
            return
        # Respaldo para archivos subidos desde la app (aun sin URL publica).
        st.caption(
            "Vista previa simplificada. Este archivo todavía no está publicado "
            "en el repositorio, por eso no se muestra con el formato original."
        )
        sheets = leer_excel(ruta)
        if len(sheets) == 1:
            nombre_hoja, df = next(iter(sheets.items()))
            st.markdown(excel_a_html(df), unsafe_allow_html=True)
        else:
            tabs = st.tabs(list(sheets.keys()))
            for tab, (nombre_hoja, df) in zip(tabs, sheets.items()):
                with tab:
                    st.markdown(excel_a_html(df), unsafe_allow_html=True)
    elif ext == ".pdf":
        data = ruta.read_bytes()
        b64 = __import__("base64").b64encode(data).decode()
        st.markdown(
            f'<iframe src="data:application/pdf;base64,{b64}" '
            f'width="100%" height="650px" style="border:none;border-radius:8px;"></iframe>',
            unsafe_allow_html=True
        )
    elif ext in (".csv",):
        df = pd.read_csv(ruta, encoding="utf-8", errors="replace")
        st.dataframe(df, use_container_width=True, hide_index=True)
    elif ext in (".txt",):
        st.code(ruta.read_text(encoding="utf-8", errors="replace"))
    else:
        st.info("Vista previa no disponible para este tipo de archivo. Usa el botón Descargar.")

# ── Header ──────────────────────────────────────────────────────────
st.markdown("""
<div class="header">
  <h1>📁 CC150701_TRI_ETE_14</h1>
  <p>Portal de Documentos · CNC</p>
</div>
""", unsafe_allow_html=True)

# ── Panel de administradora (barra lateral) ─────────────────────────
if "admin" not in st.session_state:
    st.session_state.admin = False

with st.sidebar:
    st.markdown("### 🔐 Administradora")
    if not st.session_state.admin:
        clave = st.text_input("Contraseña", type="password", key="pass_input")
        if st.button("Entrar", use_container_width=True):
            if clave == ADMIN_PASS:
                st.session_state.admin = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta")
    else:
        st.success("✓ Sesión activa")
        if st.button("Cerrar sesión", use_container_width=True):
            st.session_state.admin = False
            st.rerun()
        st.markdown("---")
        st.markdown("### ⬆️ Subir documentos")
        if not github_token():
            st.warning(
                "Falta configurar la llave de GitHub (token) en los *Secrets* "
                "de la app. Sin ella no se puede subir ni eliminar."
            )
        else:
            nuevos = st.file_uploader(
                "Selecciona uno o varios archivos",
                accept_multiple_files=True,
                label_visibility="collapsed",
                key="uploader",
            )
            if nuevos and st.button("📤 Publicar en el portal",
                                    type="primary", use_container_width=True):
                todo_ok = True
                for archi in nuevos:
                    datos = archi.getvalue()
                    ok, r = gh_subir(archi.name, datos)
                    if ok:
                        (DOCS_DIR / archi.name).write_bytes(datos)  # reflejo inmediato
                        st.success(f"✓ {archi.name}")
                    else:
                        todo_ok = False
                        detalle = (r.text[:150] if r is not None else "sin respuesta")
                        st.error(f"✗ {archi.name}: {detalle}")
                if todo_ok:
                    st.cache_data.clear()
                    st.success("Documentos publicados. Se guardaron en GitHub.")
                    st.rerun()

# ── Lista documentos ────────────────────────────────────────────────
meta = cargar_meta()

def iso_de(f):
    # Fecha del archivo: la de meta.json o, si no hay, la del commit en GitHub.
    return meta.get(f.name, {}).get("fecha", "") or fecha_en_github(f.name)

def dt_orden(iso):
    # Convierte la fecha a algo comparable; sin fecha -> queda como muy antigua.
    if not iso:
        return datetime.min
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        if dt.tzinfo:
            dt = dt.astimezone(ZONA_CO).replace(tzinfo=None)
        return dt
    except Exception:
        return datetime.min

def clave_predet(d):
    # Predeterminado: los que empiezan por "Datos" primero; el resto alfabetico.
    es_datos = 0 if d["f"].name.lower().startswith("datos") else 1
    return (es_datos, d["f"].name.lower())

docs = [
    {"f": f, "iso": iso_de(f)}
    for f in DOCS_DIR.iterdir()
    if f.is_file() and f.name != "meta.json"
]

if not docs:
    st.info("No hay documentos disponibles aún.")
else:
    col_n, col_orden = st.columns([2, 1.5])
    with col_n:
        n = len(docs)
        st.markdown(f"**{n} documento{'s' if n!=1 else ''} disponibles**")
    with col_orden:
        orden = st.selectbox(
            "🔃 Ordenar por",
            [
                "Predeterminado (Datos primero)",
                "📅 Fecha: más recientes primero",
                "📅 Fecha: más antiguos primero",
                "Nombre (A → Z)",
                "Nombre (Z → A)",
            ],
        )

    if orden == "📅 Fecha: más recientes primero":
        docs.sort(key=lambda d: dt_orden(d["iso"]), reverse=True)
    elif orden == "📅 Fecha: más antiguos primero":
        docs.sort(key=lambda d: dt_orden(d["iso"]))
    elif orden == "Nombre (A → Z)":
        docs.sort(key=lambda d: d["f"].name.lower())
    elif orden == "Nombre (Z → A)":
        docs.sort(key=lambda d: d["f"].name.lower(), reverse=True)
    else:
        docs.sort(key=clave_predet)

    st.markdown("")

    for d in docs:
        f     = d["f"]
        info  = meta.get(f.name, {})
        fecha = fmt_fecha(d["iso"])
        size  = fmt_size(info.get("size", f.stat().st_size))
        label, icon_cls = get_icon(f.name)

        with st.expander(f"{f.name}     📅 {fecha}     ·  {size}", expanded=False):
            # Fila superior: icono + info + botones
            if st.session_state.admin:
                c1, c2, c3 = st.columns([5, 1.6, 1.4])
            else:
                c1, c2 = st.columns([6, 1.6])
                c3 = None
            with c1:
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:10px;padding:4px 0">'
                    f'<div class="doc-icon {icon_cls}">{label}</div>'
                    f'<div><div class="doc-name">{f.name}</div>'
                    f'<div class="doc-meta">📅 {fecha} &nbsp;·&nbsp; {size}</div></div></div>',
                    unsafe_allow_html=True
                )
            with c2:
                st.download_button(
                    "⬇️ Descargar",
                    data=f.read_bytes(),
                    file_name=f.name,
                    key=f"dl_{f.name}",
                    use_container_width=True,
                )
            if c3 is not None:
                with c3:
                    if st.button("🗑️ Eliminar", key=f"del_{f.name}",
                                 use_container_width=True):
                        ok, r = gh_eliminar(f.name)
                        if ok:
                            f.unlink(missing_ok=True)      # reflejo inmediato
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            detalle = (r.text[:150] if r is not None else "sin respuesta")
                            st.error(f"No se pudo eliminar: {detalle}")

            # Vista previa
            st.markdown("---")
            mostrar_archivo(f)
